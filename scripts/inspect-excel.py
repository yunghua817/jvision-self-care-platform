from pathlib import Path
from openpyxl import load_workbook

path = Path.home() / "Desktop" / "JVision系統Demo清單.xlsx"
wb = load_workbook(path)
ws = wb.active
print(ws.title, ws.max_row, ws.max_column)
for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
    print(row)
print("LAST")
for row in ws.iter_rows(min_row=max(1, ws.max_row - 8), max_row=ws.max_row, values_only=True):
    print(row)
